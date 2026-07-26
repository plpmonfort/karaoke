# -*- coding: utf-8 -*-
"""
Creates Karaoke_Song_Database.xlsx from songs_seed.py.

Run this ONCE to generate the workbook. After that, edit the Excel file directly
and run build.py to push your changes into the web app.

    python3 seed_database.py            # refuses to overwrite an existing file
    python3 seed_database.py --force    # overwrites (you will lose Excel edits)
"""

import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from songs_seed import CATEGORIES

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "Karaoke_Song_Database.xlsx")

# --- House style -------------------------------------------------------------
BODY = "Arial"
HEADING_FONT = Font(name=BODY, bold=True, size=16)
SUBHEAD_FONT = Font(name=BODY, bold=True, size=16)
TABLEHEAD_FONT = Font(name=BODY, bold=True, size=10, color="FFFFFF")
CELL_FONT = Font(name=BODY, size=10)

DARK_FILL = PatternFill("solid", fgColor="1F2430")   # table header row
GREY_FILL = PatternFill("solid", fgColor="7F7F7F")   # sub-header row
BAND_FILL = PatternFill("solid", fgColor="F2F2F2")   # zebra banding

THICK_BOTTOM = Border(bottom=Side(style="medium", color="1F2430"))
THIN_BOTTOM = Border(bottom=Side(style="thin", color="7F7F7F"))
GRID = Border(
    left=Side(style="hair", color="BFBFBF"),
    right=Side(style="hair", color="BFBFBF"),
    top=Side(style="hair", color="BFBFBF"),
    bottom=Side(style="hair", color="BFBFBF"),
)

CENTRE = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)

# (header, width, alignment)
COLUMNS = [
    ("ID",         7,  CENTRE),
    ("Category",   13, CENTRE),
    ("Song Title", 34, LEFT),
    ("Artist",     28, LEFT),
    ("Year",       8,  CENTRE),
    ("Genre",      15, CENTRE),
    ("Energy",     9,  CENTRE),
    ("Difficulty", 12, CENTRE),
    ("Format",     10, CENTRE),
    ("Vocal Range", 12, CENTRE),
    ("Crowd",      8,  CENTRE),
    ("Notes",      58, LEFT),
    ("Active",     9,  CENTRE),
]

FIRST_COL = 2   # column B -- house style: leave column A empty
HEAD_ROW = 5    # table header row
DATA_ROW = 6    # first data row


def style_heading(ws, row, text, major=True):
    c = ws.cell(row=row, column=FIRST_COL, value=text)
    c.font = HEADING_FONT if major else SUBHEAD_FONT
    c.alignment = Alignment(horizontal="left", vertical="center")
    for i in range(len(COLUMNS)):
        ws.cell(row=row, column=FIRST_COL + i).border = THICK_BOTTOM if major else THIN_BOTTOM
    ws.row_dimensions[row].height = 26 if major else 22


def build_songs_sheet(wb):
    ws = wb.active
    ws.title = "Songs"
    ws.sheet_view.showGridLines = False

    style_heading(ws, 2, "KARAOKE NIGHT — MASTER SONG DATABASE", major=True)

    sub = ws.cell(row=3, column=FIRST_COL,
                  value="Edit freely, then run build.py to refresh the web app. "
                        "Set Active to N to hide a song without deleting it.")
    sub.font = Font(name=BODY, size=10, italic=True, color="595959")
    sub.alignment = Alignment(horizontal="left", vertical="center")

    # Table header
    for i, (name, width, _) in enumerate(COLUMNS):
        col = FIRST_COL + i
        ws.column_dimensions[get_column_letter(col)].width = width
        c = ws.cell(row=HEAD_ROW, column=col, value=name)
        c.font = TABLEHEAD_FONT
        c.fill = DARK_FILL
        c.alignment = CENTRE
        c.border = GRID
    ws.row_dimensions[HEAD_ROW].height = 30

    row = DATA_ROW
    sid = 1
    for cat_key, cat_label, songs in CATEGORIES:
        for (title, artist, year, genre, energy, diff, fmt, vrange, crowd, notes) in songs:
            values = [
                f"{cat_key[:1].upper()}{sid:03d}", cat_key, title, artist, year, genre,
                energy, diff, fmt, vrange, crowd, notes, "Y",
            ]
            for i, val in enumerate(values):
                c = ws.cell(row=row, column=FIRST_COL + i, value=val)
                c.font = CELL_FONT
                c.alignment = COLUMNS[i][2]
                c.border = GRID
                if sid % 2 == 0:
                    c.fill = BAND_FILL
                # Year / Energy / Crowd are counts and ratings, not measured
                # quantities -- integer format rather than the usual #,##0.0,
                # since "1,986.0" is not a year.
                if COLUMNS[i][0] in ("Year", "Energy", "Crowd"):
                    c.number_format = "0"
            ws.row_dimensions[row].height = 30
            row += 1
            sid += 1

    last = row - 1

    # Dropdowns so the sheet stays clean as you add rows (extended well past
    # the current data so new entries inherit them).
    limit = last + 500
    validations = {
        "Category": '"Warm-Up,Hype,Closer"',
        "Difficulty": '"Easy,Medium,Hard,Legend"',
        "Format": '"Solo,Duet,Group"',
        "Vocal Range": '"Low,Low-Mid,Mid,Mid-High,High,Wide"',
        "Energy": '"1,2,3,4,5"',
        "Crowd": '"1,2,3,4,5"',
        "Active": '"Y,N"',
    }
    for name, formula in validations.items():
        idx = [c[0] for c in COLUMNS].index(name)
        letter = get_column_letter(FIRST_COL + idx)
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        dv.error = f"Pick a value from the list for {name}."
        dv.errorTitle = "Not a valid option"
        ws.add_data_validation(dv)
        dv.add(f"{letter}{DATA_ROW}:{letter}{limit}")

    ws.freeze_panes = ws.cell(row=DATA_ROW, column=FIRST_COL + 2)
    ws.auto_filter.ref = (
        f"{get_column_letter(FIRST_COL)}{HEAD_ROW}:"
        f"{get_column_letter(FIRST_COL + len(COLUMNS) - 1)}{last}"
    )
    return last


def build_summary_sheet(wb, last_row):
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False
    style_heading(ws, 2, "DATABASE SUMMARY", major=True)

    for i, (name, width) in enumerate([("Category", 18), ("Phase of the Evening", 30),
                                       ("Songs", 12), ("Avg Energy", 14)]):
        col = FIRST_COL + i
        ws.column_dimensions[get_column_letter(col)].width = width
        c = ws.cell(row=4, column=col, value=name)
        c.font = TABLEHEAD_FONT
        c.fill = DARK_FILL
        c.alignment = CENTRE
        c.border = GRID
    ws.row_dimensions[4].height = 30

    cat_col = get_column_letter(FIRST_COL + 1)
    energy_col = get_column_letter(FIRST_COL + 6)
    rng_cat = f"Songs!${cat_col}${DATA_ROW}:${cat_col}${last_row}"
    rng_energy = f"Songs!${energy_col}${DATA_ROW}:${energy_col}${last_row}"

    row = 5
    for cat_key, cat_label, _ in CATEGORIES:
        cells = [
            (cat_key, None),
            (cat_label, None),
            (f'=COUNTIF({rng_cat},"{cat_key}")', "0"),
            (f'=ROUND(AVERAGEIF({rng_cat},"{cat_key}",{rng_energy}),1)', "#,##0.0"),
        ]
        for i, (val, fmt) in enumerate(cells):
            c = ws.cell(row=row, column=FIRST_COL + i, value=val)
            c.font = CELL_FONT
            c.alignment = CENTRE
            c.border = GRID
            if fmt:
                c.number_format = fmt
                c.fill = PatternFill("solid", fgColor="F2F2F2")  # formula cells
        ws.row_dimensions[row].height = 22
        row += 1

    # Total
    for i, (val, fmt) in enumerate([
        ("TOTAL", None), ("", None),
        (f"=SUM({get_column_letter(FIRST_COL+2)}5:{get_column_letter(FIRST_COL+2)}{row-1})", "0"),
        (f'=ROUND(AVERAGE({rng_energy}),1)', "#,##0.0"),
    ]):
        c = ws.cell(row=row, column=FIRST_COL + i, value=val)
        c.font = Font(name=BODY, size=10, bold=True)
        c.alignment = CENTRE
        c.border = Border(top=Side(style="medium", color="1F2430"))
        if fmt:
            c.number_format = fmt
            c.fill = PatternFill("solid", fgColor="F2F2F2")


def build_readme_sheet(wb):
    ws = wb.create_sheet("How to use")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["B"].width = 4
    ws.column_dimensions["C"].width = 110

    style_heading(ws, 2, "HOW TO USE THIS WORKBOOK", major=True)

    lines = [
        ("h", "Adding your own songs"),
        ("p", "1.  Go to the Songs sheet and type a new row directly under the last one."),
        ("p", "2.  Category, Difficulty, Format, Vocal Range, Energy, Crowd and Active are dropdowns — just pick."),
        ("p", "3.  ID can be anything unique, or left blank; build.py will fill in a fallback."),
        ("p", "4.  Save the file, then run:   python3 build.py"),
        ("p", "5.  Commit songs.js and index.html to GitHub. The web app updates automatically."),
        ("", ""),
        ("h", "What the columns mean"),
        ("p", "Category      Warm-Up = start of the evening · Hype = the middle · Closer = the end."),
        ("p", "Energy        1-5. How much the song lifts the room."),
        ("p", "Difficulty    Easy · Medium · Hard · Legend. Legend means bring your whole chest."),
        ("p", "Format        Solo · Duet · Group."),
        ("p", "Vocal Range   Roughly where the song sits: Low · Low-Mid · Mid · Mid-High · High · Wide."),
        ("p", "Crowd         1-5. How likely the room is to sing along with you."),
        ("p", "Notes         Free text. This is what shows on the back of the card in the app."),
        ("p", "Active        Y or N. Set to N to park a song without deleting it."),
        ("", ""),
        ("h", "Rebuilding from scratch"),
        ("p", "seed_database.py regenerates this workbook from songs_seed.py and will erase your edits."),
        ("p", "It refuses to run if the workbook already exists unless you pass --force."),
    ]

    row = 4
    for kind, text in lines:
        c = ws.cell(row=row, column=3, value=text)
        if kind == "h":
            c.font = Font(name=BODY, bold=True, size=11)
            ws.row_dimensions[row].height = 24
        else:
            c.font = CELL_FONT
            ws.row_dimensions[row].height = 18
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        row += 1


def main():
    if os.path.exists(OUT) and "--force" not in sys.argv:
        print(f"'{os.path.basename(OUT)}' already exists.")
        print("Refusing to overwrite your edits. Re-run with --force if you really mean it.")
        return 1

    wb = Workbook()
    last = build_songs_sheet(wb)
    build_summary_sheet(wb, last)
    build_readme_sheet(wb)
    wb.save(OUT)

    total = sum(len(s) for _, _, s in CATEGORIES)
    print(f"Wrote {os.path.basename(OUT)} — {total} songs")
    for key, label, songs in CATEGORIES:
        print(f"  {key:<9} ({label:<24}) {len(songs):>3}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
